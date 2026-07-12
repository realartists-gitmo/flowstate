#!/usr/bin/env python3
"""Build an Excel sheet of the remaining GENUINE silent errors (the ~1.9% real-fixable set) on the
full 6105-cite labelled corpus, after Sonnet gold-cleaning. Columns: raw input, raw model output,
harness output, plus the corrected (clean) answer and a category, so sol has what it needs to solve.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

BASE = "/home/adam/Projects/flowstate"
SCORED = "/tmp/claude-1000/final_scored.jsonl"
RAWS = f"{BASE}/datasets/citation_finetune/labelled_raws.jsonl"
CLEAN = "/tmp/claude-1000/clean_all.jsonl"
OUT = f"{BASE}/citation_genuine_silents.xlsx"

raws = {json.loads(l)["id"]: json.loads(l)["raw"] for l in open(RAWS)}
clean = {json.loads(l)["id"]: json.loads(l) for l in open(CLEAN)}


def is_real(r):
    """Return (category, clean_surnames) if this silent is a GENUINE repair-fixable error, else None.
    Excludes train-gold errors, org-as-blank (orgs go to publication), particle-keeping, ambiguous."""
    if not (r.get("passed") and not r.get("correct")):
        return None
    model = set(r.get("model_surnames", []))
    cl = clean.get(r["id"])
    if cl is None:
        return None  # confident gold-error bucket (not relabeled)
    g = set(cl["surnames"])
    typ = cl.get("type", "person")
    if typ == "org":
        return None  # org -> publication field, author blank => model empty is correct
    if model == g:
        return None  # train gold was wrong
    if typ == "none" and not model:
        return None
    if typ == "ambiguous":
        return None
    miss, extra = g - model, model - g
    if any(a != b and (a.endswith(b) or b.endswith(a)) and min(len(a), len(b)) >= 4
           for a in miss for b in extra):
        return None  # particle-keeping in clean label => model's particle-drop is correct
    cat = "over-extract" if (extra and not miss) else "person-miss"
    return cat, sorted(g)


rows = []
for l in open(SCORED):
    r = json.loads(l)
    res = is_real(r)
    if res is None:
        continue
    cat, gold = res
    authors = (r.get("final_obj") or {}).get("authors") or []
    harness = "; ".join(
        f"{a.get('surname','')} ({a.get('name','')})" if a.get("name") else a.get("surname", "")
        for a in authors
    ) or "(no authors)"
    rows.append({
        "raw_input": r["input"].replace("parse citation:", "").strip(),
        "raw_output": raws.get(r["id"], ""),
        "harness_output": harness,
        "correct_answer": ", ".join(gold) if gold else "(none)",
        "category": cat,
        "id": r["id"],
    })

rows.sort(key=lambda x: x["category"])
wb = Workbook()
ws = wb.active
ws.title = "genuine silents"
headers = ["Raw input (fullcite)", "Raw model output", "Harness output",
           "Correct answer (Sonnet-cleaned)", "Category", "id"]
ws.append(headers)
for c in ws[1]:
    c.font = Font(bold=True)
    c.alignment = Alignment(vertical="top")
for row in rows:
    ws.append([row["raw_input"], row["raw_output"], row["harness_output"],
               row["correct_answer"], row["category"], row["id"]])
widths = [70, 70, 40, 30, 14, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w
for r_ in ws.iter_rows(min_row=2):
    for c in r_:
        c.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"
wb.save(OUT)
print(f"wrote {len(rows)} genuine silent rows -> {OUT}")
from collections import Counter
print(Counter(r["category"] for r in rows))
