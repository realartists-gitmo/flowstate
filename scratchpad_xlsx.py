import json
import openpyxl
from openpyxl.styles import Font, Alignment

# load raw decodes
raws = {}
for l in open('scratchpad_raws_all.jsonl'):
    if l.strip():
        r = json.loads(l); raws[r['id']] = r
# load scored (has final_obj)
rows = []
for l in open('scratchpad_new_all.jsonl'):
    r = json.loads(l)
    if r['passed'] and not r['correct']:  # silent failures
        rows.append(r)
# sort by id
rows.sort(key=lambda r: r['id'])

def clean_raw_out(s):
    # the model emits braces as <unk>; show a readable JSON-ish form
    return (s or '').replace('<unk>', '{').strip()

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Silent failures"
headers = ["Raw input (cite)", "Raw model output", "Final harness output"]
ws.append(headers)
for c in range(1,4):
    ws.cell(row=1, column=c).font = Font(bold=True)
for r in rows:
    inp = r['input'].replace('parse citation:', '').strip()
    raw_out = clean_raw_out(raws.get(r['id'], {}).get('raw', ''))
    final = json.dumps(r.get('final_obj'), ensure_ascii=False, indent=None)
    ws.append([inp, raw_out, final])
# formatting: widths + wrap
ws.column_dimensions['A'].width = 70
ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 70
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
ws.freeze_panes = "A2"
wb.save('citation_silent_failures.xlsx')
print(f"wrote citation_silent_failures.xlsx with {len(rows)} silent failures")
